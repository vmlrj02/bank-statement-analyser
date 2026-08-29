"""The workbook's STRUCTURE is the product — pin it.

publish.write_workbook renders the customer's own file, Output_Template-2.xlsx,
sheet for sheet: nineteen named tabs in a fixed ORDER with fixed headers. Their
analysts read it tab by tab against their template, so a renamed sheet, a
reordered one or a shifted column is a defect to them even when every number is
right ("Xns sheet is missing" was a real bug report). Anything we add beyond
the template is appended after it, never interleaved.

Only the sanitiser had a test before this; a rewrite could silently drop a
sheet, transpose the Summary grid, or stop writing rows, and nothing would fail
until a customer opened the file.
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..",
                                "backend", "processor"))

from openpyxl import load_workbook  # noqa: E402

from bsa.models import JobResult, StatementMeta, Txn, ValidationReport  # noqa: E402
from bsa.publish import (GROUPED_HEADERS, SUMMARY_ROWS, TEMPLATE_SHEETS,  # noqa: E402
                         XN_HEADERS, write_workbook)


def _txn(desc, amount, balance, category, date, party="ACME TRADERS"):
    return Txn(date=date, cheque_no="", description=desc,
               amount=amount, balance=balance, mode="neft",
               counterparty=party, category=category,
               account_no="XX1234", bank="HDFC Bank", source_file="s.pdf")


def _result():
    meta = StatementMeta(bank="HDFC Bank", layout="hdfc", account_no="XX1234",
                         account_name="M/S STEEL", period_from="2026-01-01",
                         period_to="2026-01-31", source_file="s.pdf")
    txns = [
        _txn("CASH DEP BRANCH", 500.0, 1400.0, "cash deposit", "2026-01-05"),
        _txn("ACH D- BAJAJ FINANCE EMI", -200.0, 1200.0,
             "EMI transaction", "2026-01-06"),
        _txn("NEFT TO ACME TRADERS", -100.0, 1100.0,
             "Regular debit", "2026-01-07"),
        # A credit that is NOT business turnover, and an inward bounce charge,
        # so the Summary's rate rows have something to divide.
        _txn("INT CR", 25.0, 1125.0, "Interest received", "2026-01-15"),
        _txn("ECS RETURN CHRG", -50.0, 1075.0,
             "inward bounce penal charges", "2026-01-25"),
    ]
    return JobResult(meta=meta, txns=txns,
                     validation=ValidationReport(status="passed",
                                                 checked_rows=5, issues=[]))


def _load(tmp_path):
    out = str(tmp_path / "analysis.xlsx")
    write_workbook(_result(), out)
    return load_workbook(out)


# --- the template contract --------------------------------------------------

def test_the_first_nineteen_sheets_are_the_template_in_order(tmp_path):
    wb = _load(tmp_path)
    assert wb.sheetnames[:len(TEMPLATE_SHEETS)] == TEMPLATE_SHEETS


def test_our_extra_sheets_come_after_the_template_never_inside_it(tmp_path):
    wb = _load(tmp_path)
    extra = wb.sheetnames[len(TEMPLATE_SHEETS):]
    assert "Credit Assessment" in extra
    # The lead sheet moved deliberately: the customer's Summary is tab 1.
    assert wb.sheetnames[0] == "Summary"


def test_transaction_sheets_use_the_templates_seven_columns(tmp_path):
    wb = _load(tmp_path)
    for name in ("Xns", "Cash Deposit Xns", "EMI Xns", "SundayXns"):
        got = [c.value for c in wb[name][1]]
        assert got == XN_HEADERS, f"{name} header drifted: {got}"
    # The two grouped sheets carry the party group first and nothing else new.
    assert [c.value for c in wb["Regular Debits"][1]] == GROUPED_HEADERS


def test_rows_route_to_their_category_sheets(tmp_path):
    wb = _load(tmp_path)
    assert wb["Cash Deposit Xns"].max_row == 2       # header + the deposit
    assert wb["EMI Xns"].max_row == 2                # header + the EMI
    # Description is column 4 of the template's seven.
    assert wb["EMI Xns"].cell(row=2, column=4).value == "ACH D- BAJAJ FINANCE EMI"
    rd = wb["Regular Debits"]
    assert rd.cell(row=2, column=1).value == "ACME TRADERS"   # Group
    assert rd.cell(row=2, column=5).value == "NEFT TO ACME TRADERS"
    # A tag the master gives no destination still reaches a sheet of its own.
    assert wb["Other Xns"].max_row == 2


def test_every_transaction_appears_on_the_all_transactions_sheet(tmp_path):
    wb = _load(tmp_path)
    ws = wb["Xns"]
    assert ws.max_row == 1 + 5
    assert ws.cell(row=2, column=4).value == "CASH DEP BRANCH"


def test_serial_numbers_restart_on_every_sheet(tmp_path):
    wb = _load(tmp_path)
    # The EMI is the second transaction overall but the first on its own sheet.
    assert wb["EMI Xns"].cell(row=2, column=1).value == 1
    assert wb["Xns"].cell(row=3, column=1).value == 2


# --- the two grids ----------------------------------------------------------

def test_summary_is_the_templates_identity_block_and_month_grid(tmp_path):
    wb = _load(tmp_path)
    ws = wb["Summary"]
    assert ws["A1"].value == "Summary Info"
    assert ws["A2"].value == "Account Holder" and ws["B2"].value == "M/S STEEL"
    assert ws["A5"].value == "Account Number" and ws["B5"].value == "XX1234"
    assert ws["A9"].value == "Monthwise Details"
    # Months run ACROSS as columns, with a Total/Avg at the end.
    assert ws["A10"].value == "Item"
    assert ws["B10"].value == "Jan-2026"
    assert ws.cell(row=10, column=3).value == "Total/Avg"
    # Every row of the template's block is present, in the template's order.
    got = [ws.cell(row=11 + i, column=1).value for i in range(len(SUMMARY_ROWS))]
    assert got == SUMMARY_ROWS


def test_summary_bounce_rate_divides_by_the_masters_denominator(tmp_path):
    wb = _load(tmp_path)
    ws = wb["Summary"]
    rows = {ws.cell(row=11 + i, column=1).value: 11 + i
            for i in range(len(SUMMARY_ROWS))}
    # One inward bounce; payments issued = the EMI + the regular debit (the
    # master's list), so the rate is 1/2. Stored as a FRACTION, because the
    # template formats this row "0.00%" and Excel's percent format multiplies
    # by 100 — 50.0 in that cell would render as 5000.00%.
    rate = ws.cell(row=rows["Inward Payment Return (%)"], column=2)
    assert ws.cell(row=rows["No of Inward Bounces"], column=2).value == 1
    assert ws.cell(row=rows["Number of Payments Issued"], column=2).value == 2
    assert rate.value == 0.5
    assert rate.number_format == "0.00%"


def test_eod_balances_is_a_day_by_month_grid(tmp_path):
    wb = _load(tmp_path)
    ws = wb["EOD Balances"]
    assert ws["A1"].value == "Day/Month"
    assert ws["B1"].value == "Jan-2026"
    # Row 2 is day 1 ... row 32 is day 31, so day 5 is row 6.
    assert ws.cell(row=6, column=1).value == 5
    assert ws.cell(row=6, column=2).value == 1400.0
    # The balance carries forward into a day with no transaction: nothing moves
    # on the 8th, so it still shows the 7th's closing 1100.
    assert ws.cell(row=9, column=1).value == 8
    assert ws.cell(row=9, column=2).value == 1100.0
    assert ws.max_row == 32


def test_avg_balances_carries_the_templates_columns(tmp_path):
    wb = _load(tmp_path)
    got = [c.value for c in wb["Avg Balances"][1]]
    assert got[:6] == ["Month", "1st", "10th", "15th", "25th",
                       "Average Balance of 1st, 10th , 15th & 25th"]
    assert got[6:9] == ["Inflow", "Outflow", "Net Flow"]
    assert wb["Avg Balances"].cell(row=2, column=1).value == "Jan-2026"


def test_top_10_party_sheets_follow_the_templates_shape(tmp_path):
    import datetime

    wb = _load(tmp_path)
    ws = wb["Top 10 Party Credits"]
    # Column A is empty on this sheet in the template; the title sits in B.
    assert ws["A1"].value is None
    assert ws["B1"].value == "Monthwise Top 10 Party Credits"
    # The month is a real DATE merged across B:C and shown as mmm-yyyy, as in
    # the template — so it sorts and filters as a date, not as text.
    assert ws["B2"].value == datetime.datetime(2026, 1, 1)
    assert ws["B2"].number_format == "mmm\\-yy"   # the template's own format
    assert "B2:C2" in [str(r) for r in ws.merged_cells.ranges]
    assert [ws["B3"].value, ws["C3"].value] == ["Party", "Amount"]
    assert ws["B4"].value == "ACME TRADERS"


def test_top_10_party_months_sit_on_the_templates_fixed_stride(tmp_path):
    """Every month block is month row + header + exactly ten party rows,
    whether or not there are ten parties — so months begin at B2, B14, B26 …
    just as the template does. A variable stride looks identical but breaks
    any formula or eye-scan anchored to a row."""
    import datetime

    meta = StatementMeta(bank="HDFC Bank", layout="hdfc", account_no="XX1234",
                         account_name="M/S STEEL", period_from="2026-01-01",
                         period_to="2026-02-28", source_file="s.pdf")
    # Two months, each with a single credit party — far fewer than ten.
    txns = [_txn("NEFT IN", 500.0, 1400.0, "Regular credit", "2026-01-05",
                 party="ACME TRADERS"),
            _txn("NEFT IN", 700.0, 2100.0, "Regular credit", "2026-02-05",
                 party="BETA STEELS")]
    out = str(tmp_path / "stride.xlsx")
    write_workbook(JobResult(meta=meta, txns=txns,
                             validation=ValidationReport(status="passed",
                                                         checked_rows=2,
                                                         issues=[])), out)
    ws = load_workbook(out)["Top 10 Party Credits"]
    # Most recent month first, and the second block starts twelve rows later.
    assert ws["B2"].value == datetime.datetime(2026, 2, 1)
    assert ws["B4"].value == "BETA STEELS"
    assert ws["B14"].value == datetime.datetime(2026, 1, 1)
    assert [ws["B15"].value, ws["C15"].value] == ["Party", "Amount"]
    assert ws["B16"].value == "ACME TRADERS"


def test_annual_top_10_sheets_carry_the_templates_titles(tmp_path):
    wb = _load(tmp_path)
    annual = wb["Top 10 Credits (Annual)"]
    assert annual["A1"].value == "Top 10 Funds Received(Party Wise)"
    assert [annual["A2"].value, annual["B2"].value] == ["Description", "Amount"]
    debits = wb["Top 10 Debits (Annual)"]
    assert debits["A1"].value == "Top 10 Funds Remittance(Party Wise)"


# --- formatting: the template governs colour and format, not just structure --
#
# The customer's instruction is that the output should look exactly like
# Output_Template-2.xlsx — "every formatting colour etc". That file cannot be
# committed, so the values it dictates are pinned here instead: they were read
# out of it with openpyxl and a generated workbook was diffed against it cell
# by cell until nothing differed. If one of these changes, the output has
# drifted from the customer's file.

def test_transaction_header_wears_the_templates_navy_band(tmp_path):
    ws = _load(tmp_path)["ECS Xns"]
    head = ws["A1"]
    assert head.fill.start_color.rgb == "FF003366"
    assert head.font.name == "Arial" and head.font.b is True
    assert head.font.color.rgb == "FFFFFFFF"
    assert head.font.sz == 10
    # Frozen header, so the band stays put while an analyst scrolls.
    assert ws.freeze_panes == "A2"


def test_transaction_rows_band_on_odd_rows_only(tmp_path):
    ws = _load(tmp_path)["Xns"]
    # The template leaves row 2 unfilled and fills row 3, alternating from
    # there. Getting this inverted looks fine in isolation and wrong beside
    # their file.
    assert ws["A2"].fill.fill_type is None
    assert ws["A3"].fill.start_color.rgb == "FFCCFFFF"


def test_grouped_sheets_have_the_navy_band_but_no_banding(tmp_path):
    ws = _load(tmp_path)["Regular Credits"]
    assert ws["A1"].fill.start_color.rgb == "FF003366"
    # Regular Credits/Debits carry no row fill at all in the template.
    assert ws["A2"].fill.fill_type is None
    assert ws["A3"].fill.fill_type is None


def test_money_and_date_columns_use_the_templates_formats(tmp_path):
    # "Xns" rather than a category sheet, because only rows that exist are
    # styled — an empty sheet is just its header band.
    ws = _load(tmp_path)["Xns"]
    assert ws["B2"].number_format == "[$-409]dd\\-mmm\\-yy"
    # Amount and Balance show negatives in red parentheses; Category does not.
    assert ws["E2"].number_format == "#,##0.00_);[Red](#,##0.00)"
    assert ws["F2"].number_format == "#,##0.00"
    assert ws["G2"].number_format == "#,##0.00_);[Red](#,##0.00)"


def test_summary_and_grid_sheets_carry_their_own_header_colours(tmp_path):
    wb = _load(tmp_path)
    # Each grid sheet has a different header colour in the template.
    assert wb["EOD Balances"]["A1"].fill.start_color.rgb == "FFFFFF00"
    assert wb["EOD Balances"]["B1"].fill.start_color.rgb == "FFFBD4B4"
    assert wb["Avg Balances"]["A1"].fill.start_color.rgb == "FF333399"
    assert wb["Top 10 Credits (Annual)"]["A2"].fill.start_color.rgb == "FFBDD7EE"
    summary = wb["Summary"]
    assert summary["A2"].fill.start_color.rgb == "FFFDE9D9"   # identity label
    assert summary["B2"].fill.start_color.rgb == "FFCCFFFF"   # identity value
    assert summary["A11"].fill.start_color.rgb == "FFFFFF99"  # metric label
    assert summary["B11"].fill.start_color.rgb == "FFCCCCFF"  # metric value


def test_template_column_widths_are_applied(tmp_path):
    ws = _load(tmp_path)["ECS Xns"]
    # Description is the wide column; without this the sheet reads as a
    # column of "####" and truncated narration.
    assert round(ws.column_dimensions["D"].width, 2) == 46.71
    assert round(ws.column_dimensions["A"].width, 2) == 8.29
