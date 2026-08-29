"""Stage 6 — Publish: render outputs.

- transactions.csv : flat file (Sl. No., Date, Cheque No., Description,
  Amount in accounting format, Category, Category Detail, Party, Balance)
- analysis.xlsx    : the customer's own Output_Template workbook, sheet for
  sheet — nineteen tabs in the template's exact order and headers, with our
  own analysis (Credit Assessment, category totals) appended after them
- statement.json   : full internal records + validation report
"""
from __future__ import annotations

import csv
import json
import os
import re
from collections import defaultdict
from dataclasses import asdict
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font

from .categorize import category_detail
from .credit_summary import credit_summary
from .integrity import account_integrity
from .models import JobResult, Txn

_CS_LABELS = [
    ("months", "Months covered", "int"),
    ("avg_monthly_credits", "Avg monthly credits (all inflows)", "money"),
    ("avg_monthly_business_credits", "Avg monthly BUSINESS credits (turnover)", "money"),
    ("business_credits", "Business credits (total)", "money"),
    ("business_credit_share_pct", "Business credits as % of all credits", "pct"),
    ("avg_monthly_debits", "Avg monthly debits", "money"),
    ("total_credits", "Total credits", "money"),
    ("total_debits", "Total debits", "money"),
    ("avg_balance", "Average balance", "money"),
    ("min_balance", "Minimum balance", "money"),
    ("closing_balance", "Closing balance", "money"),
    ("cash_intensity_pct", "Cash intensity (% of turnover)", "pct"),
    ("emi_outflow_monthly", "EMI / interest outflow (monthly)", "money"),
    ("bounce_count", "Bounce / return events", "int"),
    ("penal_charges", "Penal charges", "money"),
    ("loan_disbursals", "Loan disbursals received", "money"),
    ("related_party_credit_pct", "Related-party share of credits", "pct"),
    ("distinct_credit_parties", "Distinct credit counterparties", "int"),
    ("top_party_share_pct", "Top counterparty share of credits", "pct"),
    ("turnover_trend", "Turnover trend", "text"),
    ("balance_stability_cv", "Balance stability (variation)", "num"),
    ("monthly_surplus", "Monthly surplus (net inflow)", "money"),
    ("servicing_coverage", "Debt-service coverage (turnover ÷ EMI)", "num"),
]



def _xl(v):
    """Strip characters openpyxl refuses (control chars a doctored or badly
    encoded narration can carry) — one bad cell must not kill the merge."""
    return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v


def _append(ws, row):
    ws.append([_xl(c) for c in row])

def party_key(name: str) -> str:
    """A fuzzy key that groups truncated variants of one party — the bank
    prints "MARSCONSTRUCTI" on one row and "MARSCONSTRUCTION" on another, and
    both must land in the same group. The first 12 alphanumerics collapse both
    to "MARSCONSTRUC"."""
    return re.sub(r"[^A-Za-z0-9]", "", name or "").upper()[:12]

# taxonomy tag -> destination sheet (from the "Banking extraction data
# labeling" master, column "Destination sheet in template"). A tag the master
# leaves blank has no sheet of its own: interest, investment returns and
# refunds are reported in "Xns" with their Category, exactly as the master
# specifies, and are additionally collected in the non-template "Other Xns"
# sheet so nothing is only findable by scrolling 8,000 rows.
DESTINATION_SHEETS = {
    "EMI transaction": "EMI Xns",
    "ECS transaction": "ECS Xns",
    "cash deposit": "Cash Deposit Xns",
    "cash withdrawal": "Cash Withdrawal Xns",
    "Salary paid": "Salary Paid Xns",
    "Salary credited": "Salary Paid Xns",
    "Loan amount disbursal": "Loan Disbursed Xns",
    "inward bounce penal charges": "Bounced-Penal Xns",
    "Outward Bounced Xns": "Outward Bounced Xns",
    "other penal charges": "Bounced-Penal Xns",
    "Regular credit": "Regular Credits",
    "Regular debit": "Regular Debits",
    "Related party credit": "Regular Credits",
    "Related party debit": "Regular Debits",
    "Interest received": "Other Xns",
    "Interest payments": "Other Xns",
    "Investment return credited": "Other Xns",
    "return / refund": "Other Xns",
}
# The "Group" sheets carry a leading party-group column; the others do not.
GROUPED_SHEETS = {"Regular Credits", "Regular Debits"}

# ---------------------------------------------------------------------------
# THE TEMPLATE CONTRACT — Output_Template-2.xlsx, sheet for sheet.
#
# This workbook is read by the customer's own analysts, who go down their
# template tab by tab; a renamed sheet, a reordered one or a moved column is a
# defect to them even when the numbers are right ("Xns sheet is missing" was a
# real bug report). So the nineteen names, their ORDER and their headers are
# copied from the template verbatim, down to the trailing space in "Sl. No. ".
# Anything we add beyond the template is appended AFTER these, never
# interleaved, so a template-driven reader finds every sheet where it expects.
# ---------------------------------------------------------------------------
TEMPLATE_SHEETS = [
    "Summary", "EOD Balances", "ECS Xns", "Cash Deposit Xns",
    "Cash Withdrawal Xns", "Top 10 Party Credits", "Top 10 Party Debits",
    "Loan Disbursed Xns", "EMI Xns", "Salary Paid Xns",
    "Top 10 Credits (Annual)", "Top 10 Debits (Annual)", "Avg Balances",
    "Regular Credits", "Regular Debits", "Outward Bounced Xns",
    "Bounced-Penal Xns", "Xns", "SundayXns",
]
# Transaction sheets in the template carry seven columns and no more. There is
# no Account/Bank column because a report is per ACCOUNT (gotcha 11), and the
# template has no Party column of its own — the party survives as the "Group"
# column on Regular Credits/Debits, which is where the master put it.
XN_HEADERS = ["Sl. No. ", "Date", "Cheque No.", "Description", "Amount",
              "Category", "Balance"]
GROUPED_HEADERS = ["Group"] + XN_HEADERS
# The template's transaction sheets, in the order they appear above.
XN_SHEETS = ["ECS Xns", "Cash Deposit Xns", "Cash Withdrawal Xns",
             "Loan Disbursed Xns", "EMI Xns", "Salary Paid Xns",
             "Regular Credits", "Regular Debits", "Outward Bounced Xns",
             "Bounced-Penal Xns", "Xns", "SundayXns", "Other Xns"]

# The rows of the template's "Monthwise Details" block, in the template's order.
SUMMARY_ROWS = [
    "Balance on 5th",
    "Balance on 15th",
    "Balance on 25th",
    "Average Balance (of 5th, 15th and 25th)",
    "Average Balance (month)",
    "Credit Instances No. of times (As per Statement)",
    "Total Credit Amount (As per Statement)",
    "Debit Instances No. of times (As per Statement)",
    "Total Debit Amount (As per Statement)",
    "No of Inward Bounces",
    "Number of Payments Issued",
    "Inward Payment Return (%)",
    "No of Outward Bounces",
    "Number of Payments Deposited",
    "Outward Cheque Return (%)",
]

# The CSV keeps the fuller internal shape — it is ours, not the customer's, and
# the party/category-detail columns are what the reviewer harness reads.
HEADERS = ["Sl. No.", "Date", "Account", "Bank", "Cheque No.", "Description",
           "Amount", "Category", "Category Detail", "Party", "Balance"]


def _fmt_amount(a: float) -> str:
    s = f"{abs(a):,.2f}"
    return f"({s})" if a < 0 else s


def _fmt_date(iso: str) -> str:
    # dd-mm-yyyy throughout, per the review spec (ID3).
    y, m, d = iso.split("-")
    return f"{d}-{m}-{y}"


def _row(i: int, t: Txn) -> list:
    return [i, _fmt_date(t.date), t.account_no, t.bank, t.cheque_no,
            t.description, _fmt_amount(t.amount), t.category, category_detail(t),
            t.counterparty or "unknown party", f"{t.balance:,.2f}"]


def write_csv(result: JobResult, path: str) -> None:
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for i, t in enumerate(result.txns, start=1):
            w.writerow(_row(i, t))


def write_json(result: JobResult, path: str) -> None:
    with open(path, "w") as f:
        json.dump({
            "meta": asdict(result.meta),
            "validation": result.validation.to_dict(),
            "transactions": [asdict(t) for t in result.txns],
        }, f, indent=1)


def _eod_balances(txns: list[Txn]) -> list[tuple[str, str, float]]:
    """(account, date, balance) carried forward — one series per account.

    A single series across several accounts would be meaningless, so each
    account gets its own run over its own date range.
    """
    if not txns:
        return []
    by_account: dict[str, dict[str, float]] = defaultdict(dict)
    labels: dict[str, str] = {}
    for t in txns:                       # last balance of each day wins
        k = f"{t.bank}|{t.account_no}"
        by_account[k][t.date] = t.balance
        labels[k] = t.account_no or t.bank or "—"

    out: list[tuple[str, str, float]] = []
    for k, by_day in by_account.items():
        d = date.fromisoformat(min(by_day))
        d1 = date.fromisoformat(max(by_day))
        cur = None
        while d <= d1:
            cur = by_day.get(d.isoformat(), cur)
            out.append((labels[k], d.isoformat(), cur))
            d += timedelta(days=1)
    return out


def _daily_balances(txns: list[Txn]) -> dict[str, float]:
    """date -> end-of-day balance, carried forward across days with no rows.

    A report is per ACCOUNT, so one series is the whole story here; the
    multi-account variant stays in _eod_balances for the non-template sheets.
    """
    out: dict[str, float] = {}
    for _acct, d, b in _eod_balances(txns):
        if b is not None:
            out[d] = b
    return out


_MON3 = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
         "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _month_label(mk: str) -> str:
    """'2025-01' -> 'Jan-2025', the template's month heading."""
    y, m = mk.split("-")
    return f"{_MON3[int(m) - 1]}-{y}"


def _months_desc(txns: list[Txn]) -> list[str]:
    """Calendar months present, MOST RECENT FIRST — the template's column order
    (its header row reads Jan-2025, Dec-2024, ... Feb-2024)."""
    return sorted({t.date[:7] for t in txns}, reverse=True)


def _facility(result: JobResult) -> str:
    """Savings / Current / Overdraft, as far as the statement actually says.

    An overdraft or cash-credit account prints the balance as money OWED, and
    the layout has to declare that for reconciliation to work at all — so the
    flag we already carry is the one honest signal available. Anything else is
    a guess, and a guessed facility on a credit file is worse than a blank.
    """
    layout = (getattr(result.meta, "layout", "") or "").lower()
    if "od" in layout.split("_") or "overdraft" in layout or "cash_credit" in layout:
        return "Overdraft / Cash Credit"
    if "saving" in layout:
        return "Savings"
    if "current" in layout:
        return "Current"
    return ""


def _summary_sheet(ws, result: JobResult, txns: list[Txn], bold) -> None:
    """The template's Summary: an identity block, then a month-per-COLUMN grid.

    The grid is transposed relative to every other month table we produce
    (months across, metrics down) because that is how the customer's template
    is laid out and how their analysts read it.
    """
    from .categorize import (INWARD_BOUNCE, OUTWARD_BOUNCE,
                             PAYMENTS_DEPOSITED, PAYMENTS_ISSUED)

    m = result.meta
    _append(ws, ["Summary Info"])
    ws["A1"].font = bold
    _append(ws, ["Account Holder", m.account_name or ""])
    # Address is not extracted from the statement. The row exists because the
    # template has it; filling it with anything we have not read would be a
    # fabrication on a credit document.
    _append(ws, ["Address", ""])
    _append(ws, ["Bank", m.bank or ""])
    _append(ws, ["Account Number", m.account_no or ""])
    _append(ws, ["Facility", _facility(result)])
    _append(ws, [])
    _append(ws, [])
    _append(ws, ["Monthwise Details"])
    ws["A9"].font = bold

    months = _months_desc(txns)
    _append(ws, ["Item"] + [_month_label(k) for k in months] + ["Total/Avg"])
    for c in ws[ws.max_row]:
        c.font = bold

    daily = _daily_balances(txns)

    def bal_on(mk: str, day: int):
        return daily.get(f"{mk}-{day:02d}")

    def month_bals(mk: str) -> list[float]:
        return [b for d, b in daily.items() if d[:7] == mk]

    per: dict[str, list] = {}
    for mk in months:
        rows = [t for t in txns if t.date[:7] == mk]
        cr = [t for t in rows if t.amount > 0]
        dr = [t for t in rows if t.amount < 0]
        anchors = [bal_on(mk, d) for d in (5, 15, 25)]
        present = [b for b in anchors if b is not None]
        mb = month_bals(mk)
        per[mk] = [
            anchors[0], anchors[1], anchors[2],
            round(sum(present) / len(present), 2) if present else None,
            round(sum(mb) / len(mb), 2) if mb else None,
            len(cr), round(sum(t.amount for t in cr), 2),
            len(dr), round(sum(-t.amount for t in dr), 2),
            sum(1 for t in rows if t.category == INWARD_BOUNCE),
            sum(1 for t in rows if t.category in PAYMENTS_ISSUED),
            None,                                    # inward return %, below
            sum(1 for t in rows if t.category == OUTWARD_BOUNCE),
            sum(1 for t in rows if t.category in PAYMENTS_DEPOSITED),
            None,                                    # outward return %, below
        ]
        iss, dep = per[mk][10], per[mk][13]
        per[mk][11] = round(100 * per[mk][9] / iss, 2) if iss else 0
        per[mk][14] = round(100 * per[mk][12] / dep, 2) if dep else 0

    # The last column is an average for the balance rows and a total for the
    # count/amount rows; a rate is recomputed from the totals, never averaged
    # from the monthly rates (that would weight a quiet month like a busy one).
    AVG_ROWS = {0, 1, 2, 3, 4}
    RATE_ROWS = {11: (9, 10), 14: (12, 13)}
    for i, label in enumerate(SUMMARY_ROWS):
        vals = [per[mk][i] for mk in months]
        if i in RATE_ROWS:
            num_i, den_i = RATE_ROWS[i]
            num = sum(per[mk][num_i] for mk in months)
            den = sum(per[mk][den_i] for mk in months)
            total = round(100 * num / den, 2) if den else 0
        elif i in AVG_ROWS:
            got = [v for v in vals if v is not None]
            total = round(sum(got) / len(got), 2) if got else ""
        else:
            total = round(sum(v for v in vals if v is not None), 2)
        _append(ws, [label] + [("" if v is None else v) for v in vals] + [total])
        ws[ws.max_row][0].font = bold


def _eod_sheet(ws, txns: list[Txn], bold) -> None:
    """Day-of-month down, month across — the template's EOD Balances grid."""
    months = _months_desc(txns)
    daily = _daily_balances(txns)
    _append(ws, ["Day/Month"] + [_month_label(k) for k in months])
    for c in ws[1]:
        c.font = bold
    for day in range(1, 32):
        _append(ws, [day] + [daily.get(f"{mk}-{day:02d}", "") for mk in months])


def _avg_balances_sheet(ws, txns: list[Txn], bold) -> None:
    """The template's Avg Balances: anchor-day balances plus monthly flow."""
    from .categorize import INWARD_BOUNCE, OUTWARD_BOUNCE

    daily = _daily_balances(txns)
    _append(ws, ["Month", "1st", "10th", "15th", "25th",
                 "Average Balance of 1st, 10th , 15th & 25th",
                 "Inflow", "Outflow", "Net Flow", "Inward cheque returns",
                 "Outward cheque returns", "No of credit", "No of debit"])
    for c in ws[1]:
        c.font = bold
    for mk in _months_desc(txns):
        rows = [t for t in txns if t.date[:7] == mk]
        anchors = [daily.get(f"{mk}-{d:02d}") for d in (1, 10, 15, 25)]
        got = [b for b in anchors if b is not None]
        inflow = round(sum(t.amount for t in rows if t.amount > 0), 2)
        outflow = round(sum(-t.amount for t in rows if t.amount < 0), 2)
        _append(ws, [
            _month_label(mk),
            *[("" if b is None else b) for b in anchors],
            round(sum(got) / len(got), 2) if got else "",
            inflow, outflow, round(inflow - outflow, 2),
            sum(1 for t in rows if t.category == INWARD_BOUNCE),
            sum(1 for t in rows if t.category == OUTWARD_BOUNCE),
            sum(1 for t in rows if t.amount > 0),
            sum(1 for t in rows if t.amount < 0),
        ])


def _party_month_sheet(ws, txns, want_credit, group_label, title, bold) -> None:
    """Monthwise top-10 parties: one block per calendar month, most recent
    first, starting in column B — the template leaves column A empty on these
    two sheets.

    The block geometry is the template's and is FIXED: month row, header row,
    then exactly ten party rows whether or not there are ten parties, so every
    month begins on the same 12-row stride (B2, B14, B26, ...). A variable
    stride would look the same to the eye and break any formula or eye-scan
    anchored to a row — and this is a sheet people read by scrolling to the
    month they care about. The month itself is a real date merged across B:C
    and formatted mmm-yyyy, as in the template, so it sorts and filters as a
    date rather than as the text "Feb-2026".
    """
    PER_MONTH = 10
    _append(ws, ["", title])
    ws["B1"].font = bold
    per_month: dict[str, dict[str, float]] = {}
    for t in txns:
        if (t.amount > 0) != want_credit or not t.counterparty:
            continue
        label = group_label.get(party_key(t.counterparty), t.counterparty)
        mk = t.date[:7]
        per_month.setdefault(mk, defaultdict(float))[label] += abs(t.amount)
    # The row is tracked explicitly rather than read back from ws.max_row:
    # a padding row holds no cells, so max_row does not count it and the
    # stride would silently drift on any month with fewer than ten parties.
    r = 2
    for mk in sorted(per_month, reverse=True):
        y, m = mk.split("-")
        cell = ws.cell(row=r, column=2, value=date(int(y), int(m), 1))
        cell.font = bold
        # The template's own format on this cell, verbatim: mmm-yy ("Jan-25").
        # Its Summary header uses a four-digit year instead — inconsistent in
        # their file, so each sheet matches its own.
        cell.number_format = "mmm\\-yy"
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for col, val in ((2, "Party"), (3, "Amount")):
            h = ws.cell(row=r + 1, column=col, value=val)
            h.font = bold
        ranked = sorted(per_month[mk].items(), key=lambda kv: -kv[1])[:PER_MONTH]
        for i, (party, amt) in enumerate(ranked):
            ws.cell(row=r + 2 + i, column=2, value=_xl(party))
            ws.cell(row=r + 2 + i, column=3, value=round(amt, 2))
        # Fixed stride: month + header + ten party slots, filled or not.
        r += 2 + PER_MONTH


def _party_annual_sheet(ws, txns, want_credit, group_label, title, bold) -> None:
    """Top 10 parties over the whole statement set — the template's annual view."""
    _append(ws, [title])
    ws["A1"].font = bold
    _append(ws, ["Description", "Amount"])
    for c in ws[2]:
        c.font = bold
    by_party: dict[str, float] = defaultdict(float)
    for t in txns:
        if (t.amount > 0) == want_credit and t.counterparty:
            by_party[group_label.get(party_key(t.counterparty),
                                     t.counterparty)] += abs(t.amount)
    for party in sorted(by_party, key=lambda p: -by_party[p])[:10]:
        _append(ws, [party, round(by_party[party], 2)])


def _xn_row(i: int, t: Txn) -> list:
    """One transaction in the template's seven columns."""
    return [i, _fmt_date(t.date), t.cheque_no, t.description,
            _fmt_amount(t.amount), t.category, f"{t.balance:,.2f}"]


def write_workbook(result: JobResult, path: str) -> None:
    # Belt-and-suspenders: normalize already scrubs control bytes at the source,
    # but a single illegal character in ANY cell makes openpyxl reject the whole
    # workbook ("cannot be used in worksheets"), failing the job at publish. So
    # scrub the meta strings again here in case a merge path re-derived them —
    # the txn fields are already clean from normalize.
    from .normalize import scrub_control
    m = result.meta
    for fld in ("account_name", "account_no", "bank", "producer", "creator",
                "pdf_created", "pdf_modified"):
        if hasattr(m, fld):
            setattr(m, fld, scrub_control(getattr(m, fld)))

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    big = Font(bold=True, size=13)
    txns = result.txns
    integ = account_integrity([result.meta], result.validation.status)

    # Create every template sheet up front, in the template's order, so the tab
    # bar matches the customer's file even when a sheet ends up empty.
    sheets = {name: wb.create_sheet(name) for name in TEMPLATE_SHEETS}

    # The fullest party name seen wins for a fuzzy group, so "MARSCONSTRUCTI"
    # and "MARSCONSTRUCTION" collapse to one party everywhere in the workbook.
    group_label: dict[str, str] = {}
    for t in txns:
        k = party_key(t.counterparty)
        if k and len(t.counterparty or "") > len(group_label.get(k, "")):
            group_label[k] = t.counterparty

    _summary_sheet(sheets["Summary"], result, txns, bold)
    _eod_sheet(sheets["EOD Balances"], txns, bold)
    _avg_balances_sheet(sheets["Avg Balances"], txns, bold)
    _party_month_sheet(sheets["Top 10 Party Credits"], txns, True, group_label,
                       "Monthwise Top 10 Party Credits", bold)
    _party_month_sheet(sheets["Top 10 Party Debits"], txns, False, group_label,
                       "Monthwise Top 10 Party Debits", bold)
    _party_annual_sheet(sheets["Top 10 Credits (Annual)"], txns, True,
                        group_label, "Top 10 Funds Received(Party Wise)", bold)
    _party_annual_sheet(sheets["Top 10 Debits (Annual)"], txns, False,
                        group_label, "Top 10 Funds Remittance(Party Wise)", bold)

    # --- transaction sheets -------------------------------------------------
    # "Other Xns" is ours, not the template's, so it is created after the
    # nineteen and holds the tags the master gives no destination sheet
    # (interest, investment returns, refunds) rather than leaving them
    # findable only by scrolling Xns.
    sheets["Other Xns"] = wb.create_sheet("Other Xns")
    for name in XN_SHEETS:
        ws = sheets[name]
        _append(ws, GROUPED_HEADERS if name in GROUPED_SHEETS else XN_HEADERS)
        for c in ws[1]:
            c.font = bold
    # Sl. No. restarts on every sheet: on a template sheet it numbers that
    # sheet's rows, which is what a reader counting cash deposits expects.
    seq: dict[str, int] = defaultdict(int)

    def put(name: str, t: Txn) -> None:
        seq[name] += 1
        row = _xn_row(seq[name], t)
        if name in GROUPED_SHEETS:
            row = [group_label.get(party_key(t.counterparty),
                                   t.counterparty or "unknown party")] + row
        _append(sheets[name], row)

    for t in txns:
        put(DESTINATION_SHEETS.get(t.category, "Other Xns"), t)
        put("Xns", t)
        if date.fromisoformat(t.date).weekday() == 6:
            put("SundayXns", t)

    # --- sheets beyond the template, appended after it ----------------------
    # Credit Assessment is the lender-facing conclusion and the thing this
    # product is actually for; it sits after the template so the customer's
    # nineteen tabs are exactly where their file says they are.
    cs = credit_summary(txns, integ, result.validation.status)
    ws = wb.create_sheet("Credit Assessment")
    _append(ws, ["Credit Assessment"])
    ws["A1"].font = big
    _append(ws, [result.meta.account_name or "", result.meta.bank or "",
                 result.meta.account_no or ""])
    _append(ws, ["Balance reconciliation", result.validation.status,
                 "Integrity", integ["assessment"]])
    from .completeness import check_completeness
    nd = sum(1 for t in txns if t.amount < 0)
    nc = sum(1 for t in txns if t.amount > 0)
    sd = sum(-t.amount for t in txns if t.amount < 0)
    sc = sum(t.amount for t in txns if t.amount > 0)
    comp = check_completeness(len(txns), nd, nc,
                              getattr(result.meta, "declared_totals", None) or {},
                              sd, sc)
    if comp.get("checked"):
        _append(ws, ["Completeness",
                     "complete" if comp["complete"] else "INCOMPLETE",
                     "; ".join(comp.get("notes", [])) or
                     f"{len(txns)} of {comp['declared']} declared"])
    _append(ws, [])
    _append(ws, ["Statement period", result.meta.period_from or "",
                 "to", result.meta.period_to or ""])
    _append(ws, [])
    _append(ws, ["Metric", "Value"])
    for c in ws[ws.max_row]:
        c.font = bold
    metrics = cs["metrics"]
    for key, label, kind in _CS_LABELS:
        v = metrics.get(key)
        if kind == "pct":
            v = f"{v}%"
        elif kind == "money" and isinstance(v, (int, float)):
            v = _fmt_amount(v)
        _append(ws, [label, v])
    # Speculative / high-risk spending, named group by group, so the sheet
    # says WHERE the money went rather than only that a flag fired.
    hr = metrics.get("high_risk_spend") or {}
    if hr:
        _append(ws, [])
        _append(ws, ["Speculative / high-risk spending", "Count", "Amount"])
        for c in ws[ws.max_row]:
            c.font = bold
        for g, v in sorted(hr.items(), key=lambda kv: -kv[1]["amount"]):
            _append(ws, [g, v["count"], _fmt_amount(-v["amount"])])
    _append(ws, [])
    _append(ws, ["Underwriting reads"])
    ws[ws.max_row][0].font = bold
    for r in cs["reads"]:
        _append(ws, ["", r])

    # Category totals and the per-account roll-up that the template's Summary
    # has no room for.
    ws = wb.create_sheet("Category Totals")
    agg: dict[str, list[float]] = defaultdict(lambda: [0, 0.0])
    for t in txns:
        agg[t.category][0] += 1
        agg[t.category][1] += t.amount
    _append(ws, ["Category", "Count", "Net Amount"])
    for c in ws[1]:
        c.font = bold
    for tag in sorted(agg, key=lambda k: -abs(agg[k][1])):
        _append(ws, [tag, agg[tag][0], round(agg[tag][1], 2)])
    _append(ws, [])
    _append(ws, ["Integrity", integ["assessment"].upper()])
    ws[ws.max_row][0].font = bold
    _append(ws, ["PDF producer", result.meta.producer or "—",
                 "Created", result.meta.pdf_created or "—",
                 "Modified", result.meta.pdf_modified or "—"])
    for flag in integ["flags"]:
        _append(ws, ["", flag])
    _append(ws, [])
    _append(ws, ["Validation", result.validation.status,
                 f"{result.validation.checked_rows} rows checked",
                 f"{len(result.validation.issues)} issues"])

    # Largest SINGLE transactions — a different question from the biggest
    # parties, and the one that finds a one-off ₹50L movement.
    for name, want_credit in (("Top 10 Credits(Consolidated)", True),
                              ("Top 10 Debits(Consolidated)", False)):
        ws = wb.create_sheet(name)
        _append(ws, ["Date", "Description", "Party", "Amount"])
        for c in ws[1]:
            c.font = bold
        ranked = sorted((t for t in txns if (t.amount > 0) == want_credit),
                        key=lambda t: -abs(t.amount))[:10]
        for t in ranked:
            _append(ws, [_fmt_date(t.date), t.description,
                         t.counterparty or "unknown party",
                         _fmt_amount(t.amount)])

    wb.save(path)


def publish(result: JobResult, out_dir: str, basename: str = "statement") -> dict[str, str]:
    os.makedirs(out_dir, exist_ok=True)
    paths = {
        "csv": os.path.join(out_dir, f"{basename}_transactions.csv"),
        "xlsx": os.path.join(out_dir, f"{basename}_analysis.xlsx"),
        "json": os.path.join(out_dir, f"{basename}.json"),
    }
    write_csv(result, paths["csv"])
    write_workbook(result, paths["xlsx"])
    write_json(result, paths["json"])
    return paths
