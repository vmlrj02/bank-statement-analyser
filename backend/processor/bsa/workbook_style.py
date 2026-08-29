"""Cell formatting copied from the customer's Output_Template-2.xlsx.

Every colour, font, border, column width and number format in this module was
READ OUT of that workbook with openpyxl — none of it was chosen by us. The
customer's instruction is that the output should look exactly like their
template, so when a styling question comes up the answer is to open the
template and copy what it does, not to pick something reasonable.

The template's visual language, for reference:

  * Arial throughout. 10pt for headers and data, 9pt in the Summary identity
    block, 12pt and 14pt for the two sheet titles.
  * Transaction sheets carry a deep-navy header band with white bold text, and
    band their data rows in pale cyan on ODD rows only (row 2 unfilled, row 3
    filled, and so on). The two grouped sheets — Regular Credits/Debits — use
    the same navy header but no banding at all.
  * Grid sheets each have their own header colour: yellow corner and peach
    months on EOD Balances, indigo on Avg Balances, ice blue on the annual
    top-10 sheets, yellow on the monthwise top-10 blocks.
  * Money is "#,##0.00_);[Red](#,##0.00)" — negatives in red parentheses —
    except where the template uses plain "#,##0.00", which is preserved
    per column rather than normalised.

Anything the product adds beyond the customer's nineteen sheets (Credit
Assessment and the extra views) is deliberately NOT styled from here: those
sheets are ours, the template says nothing about them, and dressing them in
the customer's colours would imply they are part of the contract.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- the template's palette, ARGB exactly as the file stores it -------------
NAVY = "FF003366"        # transaction-sheet header band
INDIGO = "FF333399"      # Avg Balances header band
YELLOW = "FFFFFF00"      # EOD corner cell, monthwise top-10 column heads
PEACH = "FFFBD4B4"       # EOD month headers and day column
CREAM = "FFFDE9D9"       # Summary identity labels
CYAN = "FFCCFFFF"        # Summary identity values, transaction row banding
SAND = "FFDDD9C3"        # month bands (Summary, Avg Balances, top-10)
STEEL = "FFDBE5F1"       # Summary month-band corner
LEMON = "FFFFFF99"       # Summary metric labels
PERIWINKLE = "FFCCCCFF"  # Summary metric values, Avg Balances values
SKY = "FFC6D9F0"         # EOD balance values
BLUSH = "FFFDEADA"       # monthwise top-10 party cells
PALEBLUE = "FFDCE6F2"    # monthwise top-10 amount cells
ICE = "FFBDD7EE"         # annual top-10 header
BUTTER = "FFFFF2CC"      # annual top-10 data

# --- number formats ---------------------------------------------------------
DATE = "[$-409]dd\\-mmm\\-yy"
MONTH = "[$-409]mmm\\-yy"
MONTH_PLAIN = "mmm\\-yy"
MONEY = "#,##0.00"
MONEY_RED = "#,##0.00_);[Red](#,##0.00)"
INT = "0"
PCT = "0.00%"
TEXT = "@"
GENERAL = "General"


def _f(colour: str) -> PatternFill:
    return PatternFill("solid", start_color=colour, end_color=colour)


_THIN = Side(style="thin", color="FF000000")
_MEDIUM = Side(style="medium", color="FF000000")
BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BOX_MEDIUM = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)


def _b(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN) -> Border:
    """A border with any side overridden — the template edges its grids with
    `medium` on the outside and drops some sides entirely, so the four are set
    individually rather than as one box."""
    return Border(left=left, right=right, top=top, bottom=bottom)


# The template stores an explicit black on every body font; leaving it unset
# would be theme-default black and compare unequal.
BLACK = "FF000000"
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEAD_DARK = Font(name="Arial", size=10, bold=True, color=BLACK)
BODY = Font(name="Arial", size=10, color=BLACK)
BODY_SMALL = Font(name="Arial", size=9, color=BLACK)
TITLE_14 = Font(name="Arial", size=14, color=BLACK)
TITLE_12 = Font(name="Arial", size=12, bold=True, color=BLACK)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NOWRAP = Alignment(horizontal="center", vertical="center")
MIDDLE = Alignment(vertical="center")


def _widths(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# Column widths, read from the template sheet by sheet. The last entry on each
# is the narrow spacer the template leaves just past the data.
XN_WIDTHS = {"A": 8.29, "B": 10.86, "C": 12.0, "D": 46.71, "E": 15.57,
             "F": 37.0, "G": 16.43, "H": 9.0}
GROUPED_WIDTHS = {"A": 7.29, "B": 7.43, "C": 10.86, "D": 11.43, "E": 46.71,
                  "F": 15.57, "G": 37.0, "H": 16.43, "I": 9.0}

# Per-column number formats on the transaction sheets, in template order:
# Sl. No., Date, Cheque No., Description, Amount, Category, Balance.
XN_FORMATS = [GENERAL, DATE, INT, GENERAL, MONEY_RED, MONEY, MONEY_RED]
GROUPED_FORMATS = [GENERAL] + XN_FORMATS
# The Xns tab carries an extra "Party Name" column between Category and
# Balance (see publish.PARTY_COLUMN_SHEETS), so its formats and widths take
# the party column's shape from the grouped sheets' party column.
PARTY_COL_FORMATS = XN_FORMATS[:-1] + [GENERAL, XN_FORMATS[-1]]
PARTY_COL_WIDTHS = dict(XN_WIDTHS, G=24.0, H=16.43, I=9.0)


def style_xn_sheet(ws, grouped: bool = False, party_col: bool = False) -> None:
    """A transaction sheet: navy header, thin box borders everywhere, and —
    on the ungrouped sheets only — pale cyan banding on odd data rows."""
    if grouped:
        fmts, widths = GROUPED_FORMATS, GROUPED_WIDTHS
    elif party_col:
        fmts, widths = PARTY_COL_FORMATS, PARTY_COL_WIDTHS
    else:
        fmts, widths = XN_FORMATS, XN_WIDTHS
    _widths(ws, widths)
    ws.freeze_panes = "A2"
    ncols = len(fmts)
    for i in range(ncols):
        c = ws.cell(row=1, column=i + 1)
        c.fill = _f(NAVY)
        c.font = HEAD_FONT
        c.border = BOX
        c.alignment = CENTER
        # The template carries the money format on the Amount, Category and
        # Balance header cells as well as on their data. Those are the last
        # three columns on a normal sheet, but not once the Xns tab gains a
        # Party Name column between Category and Balance — so the money header
        # cells are chosen by position from the RIGHT, skipping the party one.
        money_cols = ({ncols - 4, ncols - 3, ncols - 1} if party_col
                      else {ncols - 3, ncols - 2, ncols - 1})
        c.number_format = MONEY if i in money_cols else GENERAL
    band = _f(CYAN)
    for r in range(2, ws.max_row + 1):
        for i in range(ncols):
            c = ws.cell(row=r, column=i + 1)
            c.font = BODY
            c.border = BOX
            c.number_format = fmts[i]
            c.alignment = MIDDLE
            # The template bands the ungrouped sheets on odd rows only; the
            # grouped ones carry no fill at all.
            if not grouped and r % 2 == 1:
                c.fill = band


def style_summary(ws, n_months: int, pct_rows: set[int]) -> None:
    """The template's Summary: a 9pt identity block over a month-per-column
    grid. `pct_rows` are indices into the metric block that the template
    formats as percentages."""
    _widths(ws, {"A": 42.14, "B": 16.43, "O": 9.0})
    ws.freeze_panes = "B1"
    ws["A1"].font = TITLE_14
    ws["A1"].alignment = MIDDLE
    # Identity block: label column cream, value column cyan, both 9pt.
    for r in range(2, 7):
        lab, val = ws.cell(row=r, column=1), ws.cell(row=r, column=2)
        lab.fill, val.fill = _f(CREAM), _f(CYAN)
        lab.font = val.font = BODY_SMALL
        lab.border = val.border = BOX
        lab.alignment = Alignment(vertical="center", wrap_text=True)
        val.alignment = Alignment(horizontal="right", vertical="center")
    # Account Number is text in the template, so a long number keeps its
    # leading zeros and never renders in scientific notation.
    ws.cell(row=5, column=2).number_format = TEXT
    ws["A9"].font = HEAD_DARK

    ncols = n_months + 2                      # label + months + Total/Avg
    for i in range(ncols):                    # month band
        c = ws.cell(row=10, column=i + 1)
        c.fill = _f(STEEL if i == 0 else SAND)
        c.font = HEAD_DARK
        c.border = BOX
        c.alignment = CENTER
        if 0 < i <= n_months:
            c.number_format = MONTH
    for r in range(11, ws.max_row + 1):
        fmt = PCT if (r - 11) in pct_rows else MONEY_RED
        for i in range(ncols):
            c = ws.cell(row=r, column=i + 1)
            c.border = BOX
            if i == 0:
                c.fill, c.font = _f(LEMON), BODY
            else:
                c.fill, c.font = _f(PERIWINKLE), BODY
                c.number_format = fmt


def style_eod(ws, n_months: int) -> None:
    """EOD Balances: yellow corner, peach months across and days down."""
    _widths(ws, {"A": 10.14, "B": 17.14, "C": 18.86, "T": 9.0})
    ws.freeze_panes = "B2"
    # The grid is edged: medium along the top row and down the left column.
    corner = ws.cell(row=1, column=1)
    corner.fill, corner.font = _f(YELLOW), HEAD_DARK
    corner.border = _b(left=_MEDIUM, top=_MEDIUM)
    corner.alignment = CENTER_NOWRAP
    for i in range(1, n_months + 1):
        c = ws.cell(row=1, column=i + 1)
        c.fill, c.font = _f(PEACH), HEAD_DARK
        c.border, c.alignment = _b(top=_MEDIUM), CENTER_NOWRAP
        c.number_format = MONTH
    for r in range(2, ws.max_row + 1):
        day = ws.cell(row=r, column=1)
        day.fill, day.font = _f(PEACH), HEAD_DARK
        day.border = _b(left=_MEDIUM)
        for i in range(1, n_months + 1):
            c = ws.cell(row=r, column=i + 1)
            c.fill, c.font, c.border = _f(SKY), BODY, BOX
            c.number_format = MONEY_RED


def style_avg_balances(ws, ncols: int) -> None:
    """Avg Balances: an indigo header band with medium borders."""
    _widths(ws, {"A": 12.57, "B": 18.86, "J": 19.57, "K": 20.14, "L": 12.43,
                 "M": 14.29, "N": 8.0})
    ws.freeze_panes = "A2"
    for i in range(ncols):
        c = ws.cell(row=1, column=i + 1)
        c.fill, c.font = _f(INDIGO), HEAD_FONT
        # Only the first header cell carries a left edge; the rest run together.
        c.border = BOX_MEDIUM if i == 0 else _b(left=None, right=_MEDIUM,
                                                top=_MEDIUM, bottom=_MEDIUM)
        c.alignment = CENTER
    for r in range(2, ws.max_row + 1):
        mon = ws.cell(row=r, column=1)
        mon.fill, mon.font, mon.border = _f(SAND), HEAD_DARK, BOX
        mon.number_format = MONTH
        for i in range(1, ncols):
            c = ws.cell(row=r, column=i + 1)
            c.fill, c.font, c.border = _f(PERIWINKLE), BODY, BOX
            # The last four columns are counts in the template, not money.
            c.number_format = INT if i >= ncols - 4 else MONEY_RED


def style_party_month(ws, month_rows: list[int]) -> None:
    """Monthwise top-10: a fixed 12-row block per month starting in column B.

    `month_rows` are the sheet rows carrying each month label; the header sits
    directly below and ten party rows below that."""
    _widths(ws, {"A": 2.14, "B": 37.29, "C": 22.86, "D": 9.0})
    ws["B1"].font = TITLE_12
    ws["B1"].border = BOX
    ws["B1"].alignment = CENTER_NOWRAP
    for mr in month_rows:
        m = ws.cell(row=mr, column=2)
        m.fill, m.font, m.border = _f(SAND), HEAD_DARK, BOX
        m.number_format = MONTH_PLAIN
        m.alignment = CENTER_NOWRAP
        ws.cell(row=mr, column=3).fill = _f(SAND)
        ws.cell(row=mr, column=3).border = BOX
        # Below the month band the block has no horizontal rules between
        # cells (no top border), and the Amount column has no left edge — the
        # two columns read as one panel.
        for col in (2, 3):                    # Party / Amount header
            h = ws.cell(row=mr + 1, column=col)
            h.fill, h.font = _f(YELLOW), HEAD_DARK
            h.border = _b(top=None, left=_THIN if col == 2 else None)
            h.alignment = CENTER_NOWRAP
        for r in range(mr + 2, mr + 12):      # ten party slots
            p, a = ws.cell(row=r, column=2), ws.cell(row=r, column=3)
            p.fill, a.fill = _f(BLUSH), _f(PALEBLUE)
            p.font = a.font = BODY
            p.border = _b(top=None)
            a.border = _b(top=None, left=None)
            a.number_format = MONEY


def style_party_annual(ws) -> None:
    """Annual top-10: an ice-blue title and header over butter-yellow rows."""
    _widths(ws, {"A": 32.86, "B": 21.43, "C": 9.0})
    # The block is edged medium down its outside — left of column A, right of
    # column B — with the title's own bottom rule left thin.
    for col in (1, 2):
        outer = {"left": _MEDIUM} if col == 1 else {"right": _MEDIUM}
        t = ws.cell(row=1, column=col)
        t.fill, t.font = _f(ICE), TITLE_14
        t.border = _b(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_THIN)
        t.alignment = CENTER
        h = ws.cell(row=2, column=col)
        h.fill, h.font = _f(ICE), HEAD_DARK
        h.border, h.alignment = _b(**outer), CENTER
    for r in range(3, ws.max_row + 1):
        for col in (1, 2):
            outer = {"left": _MEDIUM} if col == 1 else {"right": _MEDIUM}
            c = ws.cell(row=r, column=col)
            c.fill, c.font, c.border = _f(BUTTER), BODY, _b(**outer)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=2).number_format = MONEY_RED
